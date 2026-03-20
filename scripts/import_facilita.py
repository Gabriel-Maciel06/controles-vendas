import openpyxl
import requests
import json
import os
from datetime import datetime, timedelta

FILE_PATH = r"C:\Users\barth\OneDrive\Desktop\FacilitaPasta\FACILITA VENDAS ATUALIZADO 09-03-2026.xlsx"
API_URL = "https://controles-vendas.onrender.com/api/import/facilita"

SELLERS_MAP = {
    "MACIEL": "default",
    "CAIO": "caio",
    "KARINE": "karine",
    "FERNANDA": "fernanda"
}

def parse_date(value):
    if not value: return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        try:
            if '-' in value:
                return datetime.strptime(value.split(' ')[0], '%Y-%m-%d')
            else:
                return datetime.strptime(value.split(' ')[0], '%d/%m/%Y')
        except:
            pass
    return None

def main():
    if not os.path.exists(FILE_PATH):
        print(f"ALERTA: Arquivo não encontrado em {FILE_PATH}.")
        return

    print(f"Lendo '{FILE_PATH}' (read_only mode)...")
    wb = openpyxl.load_workbook(FILE_PATH, read_only=True, data_only=True)
    sheet_name = "BASE FACILITA" if "BASE FACILITA" in wb.sheetnames else wb.sheetnames[0]
    sheet = wb[sheet_name]

    headers = {}
    row_iter = sheet.iter_rows(values_only=True)
    
    try:
        first_row = next(row_iter)
    except StopIteration:
        print("Planilha vazia.")
        return

    for idx, col_name in enumerate(first_row):
        if col_name:
            headers[str(col_name).strip().upper()] = idx

    print(f"Mapeamento de colunas: {list(headers.keys())}")
    
    agrupados = {v: {} for v in SELLERS_MAP.keys()}

    print("Processando e deduplicando base (isso pode levar alguns minutos)...")
    count = 0
    for row in row_iter:
        count += 1
        if count % 50000 == 0:
            print(f"Lidas {count} linhas...")

        vendedor_idx = headers.get("VENDEDOR", 7)
        vendedor = str(row[vendedor_idx]).strip().upper() if len(row) > vendedor_idx and row[vendedor_idx] else None
        
        if vendedor not in SELLERS_MAP:
            continue
        
        cod_idx = headers.get("CÓD.C", 0)
        cod_c = str(row[cod_idx]).strip() if len(row) > cod_idx else None
        if not cod_c or cod_c == 'None':
            continue

        date_idx = headers.get("DATA PEDIDO", 3)
        data_pedido = parse_date(row[date_idx]) if len(row) > date_idx else None
        
        if cod_c not in agrupados[vendedor]:
            cli_idx = headers.get("CLIENTE/FORNEC.", 1)
            sit_idx = headers.get("SITUAÇÃO", 8)
            reg_idx = headers.get("REGIÃO", 9)
            cid_idx = headers.get("CIDADE", 10)
            
            freq_idx = headers.get("FREQUENCIA COMPRA DIAS", -1)
            med_idx = headers.get("MÉDIA POR COMPRAS", -1)

            agrupados[vendedor][cod_c] = {
                "COD": cod_c,
                "CLIENTE": str(row[cli_idx]).strip() if len(row) > cli_idx else "",
                "VALOR_TOTAL": 0.0,
                "DATA_MAIS_RECENTE": data_pedido,
                "SITUACAO": str(row[sit_idx]).strip().upper() if len(row) > sit_idx and row[sit_idx] else "ATIVO",
                "REGIAO": str(row[reg_idx]).strip() if len(row) > reg_idx and row[reg_idx] else "",
                "CIDADE": str(row[cid_idx]).strip() if len(row) > cid_idx and row[cid_idx] else "",
                "FREQ_DIAS": str(row[freq_idx]).strip() if freq_idx >= 0 and len(row) > freq_idx and row[freq_idx] else "",
                "MEDIA_COMPRAS": str(row[med_idx]).strip() if med_idx >= 0 and len(row) > med_idx and row[med_idx] else ""
            }
        
        # Correção no mapeamento da Data e Frequência para a mais recente
        if data_pedido:
            atual_recente = agrupados[vendedor][cod_c]["DATA_MAIS_RECENTE"]
            if not atual_recente or data_pedido > atual_recente:
                agrupados[vendedor][cod_c]["DATA_MAIS_RECENTE"] = data_pedido
                freq_idx = headers.get("FREQUENCIA COMPRA DIAS", -1)
                med_idx = headers.get("MÉDIA POR COMPRAS", -1)
                if freq_idx >= 0 and len(row) > freq_idx: agrupados[vendedor][cod_c]["FREQ_DIAS"] = str(row[freq_idx]).strip()
                if med_idx >= 0 and len(row) > med_idx: agrupados[vendedor][cod_c]["MEDIA_COMPRAS"] = str(row[med_idx]).strip()

        try:
            val_idx = headers.get("VALOR NOTA", 2)
            if len(row) > val_idx and row[val_idx]:
                val = float(str(row[val_idx]).replace('R$', '').replace('.','').replace(',','.').strip())
                agrupados[vendedor][cod_c]["VALOR_TOTAL"] += val
        except:
            pass

    wb.close()
    
    for vendedor, base_profile in SELLERS_MAP.items():
        clientes_vendedor = list(agrupados[vendedor].values())
        total = len(clientes_vendedor)
        print(f"\n[{vendedor}] Preparando {total} clientes únicos para envio REST...")
        
        if total == 0: continue
        
        payload_customers = []
        for c in clientes_vendedor:
            situacao = c["SITUACAO"]
            origin = "Prospec" if situacao == "ATIVO" else "Inativo"
            
            data_rec_str = c["DATA_MAIS_RECENTE"].strftime("%Y-%m-%d") if c["DATA_MAIS_RECENTE"] else ""
            data_br = c["DATA_MAIS_RECENTE"].strftime("%d/%m/%Y") if c["DATA_MAIS_RECENTE"] else "Desconhecida"
            
            next_follow = ""
            if c["DATA_MAIS_RECENTE"]:
                next_follow = (c["DATA_MAIS_RECENTE"] + timedelta(days=7)).strftime("%Y-%m-%d")
                
            notas = f"Base Facilita | Freq: {c['FREQ_DIAS']}d | Ticket médio: {c['MEDIA_COMPRAS']} | Última compra: {data_br} | Valor Histórico: R$ {c['VALOR_TOTAL']:.2f}"

            cliente_final = {
                "id": f"facilita_{c['COD']}_{vendedor}",
                "profile": base_profile,
                "name": c["CLIENTE"][:100] if c["CLIENTE"] and c["CLIENTE"] != 'None' else "Sem Nome",
                "source": situacao,
                "origin": origin,
                "temperature": "Frio",
                "status": "Frio",
                "region": c["REGIAO"] if c["REGIAO"] and c["REGIAO"] != 'None' else "",
                "city": c["CIDADE"] if c["CIDADE"] and c["CIDADE"] != 'None' else "",
                "lastContactDate": data_rec_str,
                "nextFollowUp": next_follow,
                "notes": notas,
                "createdAt": datetime.now().isoformat()
            }
            payload_customers.append(cliente_final)
            
        batch_size = 100
        t_criados, t_ignorados, t_erros = 0, 0, 0
        
        import_log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "import_log.txt")
        with open(import_log_path, "a", encoding="utf-8") as f_log:
            f_log.write(f"\n--- Iniciando Importação {vendedor} ({total} registros) ---\n")
            
            for i in range(0, total, batch_size):
                lote = payload_customers[i:i+batch_size]
                body = {
                    "customers": lote,
                    "profile": "default"
                }
                
                print(f"[{vendedor}] Importando lote {i//batch_size + 1}/{(total//batch_size)+1}... ({min(i+batch_size, total)}/{total})")
                try:
                    res = requests.post(API_URL, json=body, timeout=30)
                    if res.status_code == 200:
                        data = res.json()
                        t_criados += data.get("criados", 0)
                        t_ignorados += data.get("ignorados", 0)
                        t_erros += data.get("erros", 0)
                        f_log.write(f"Lote {i//batch_size + 1}: {data}\n")
                    else:
                        print(f"Erro no lote. HTTP {res.status_code}: {res.text}")
                        f_log.write(f"ERRO HTTP {res.status_code}: {res.text}\n")
                except Exception as e:
                    print(f"Falha de conexão com a API: {e}")
                    f_log.write(f"FALHA CONEXÃO: {e}\n")
                    
        print(f"[{vendedor}] Concluído: {t_criados} criados, {t_ignorados} ignorados, {t_erros} erros limitados/preservados.")

if __name__ == "__main__":
    main()
